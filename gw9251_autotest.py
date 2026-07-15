"""
GW9251 24bit ADC characterization automation
- Auto-identifies the DAC board (gw9241) vs the test board (gw9251) by
  probing each port's firmware prompt — no config file needed
- Sets DAC voltages, writes/verifies registers, captures 1024 samples
- Pastes results into the Excel report via openpyxl (CSV backup saved too)

Usage:
    pip install pyserial openpyxl
    python gw9251_autotest.py --dut 3
"""

import argparse
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import serial
from serial.tools import list_ports

# ----------------------------- configuration -----------------------------

BAUD = 115200          # match Tera Term serial port setting
LINE_END = "\r"        # Tera Term sends CR on Enter; CRLF corrupts the next command
CMD_DELAY = 0.15       # seconds to wait after each command
READ_TIMEOUT = 2.0     # per-line read timeout
# rd streams ~11 samples/s -> 1024 samples take ~95 s. Abort only when the
# stream goes QUIET (idle), with a generous hard cap as a safety net.
# 20 s idle: a run stalled mid-stream for >5 s at sample ~513, so give
# transient DRDY hiccups room to recover before declaring the stream dead.
CAPTURE_IDLE_TIMEOUT = 20.0  # seconds without data = stream ended
CAPTURE_HARD_TIMEOUT = 300.0 # absolute max wait for a capture

DAC_HIGH = 1520000     # dac set 2 value (AINP)
DAC_LOW = 1480000      # dac set 1 value (AINN)

N_SAMPLES = 1024

EXCEL_PATH = Path("GW9251_ADC_report.xlsx")  # override with --excel
SHEET_NAME = None          # None = active sheet; override with --sheet
HEADER_ROW = 5             # row containing '... DUT#n' merged headers
MODE_ROW = 10              # row containing 'Internal Short' / 'Channel A'
DATA_START_ROW = 22        # first sample row (B22=0 ... B1045=1023)

# ----------------------------- serial helpers -----------------------------

# firmware output contains NUL bytes and ANSI escape codes (Tera Term silently
# strips these; pyserial does not) — clean before matching or parsing anything.
# note: firmware emits non-standard params like ESC[-1D, hence the '-'
ANSI_RE = re.compile(r"\x1b\[[0-9;?\-]*[A-Za-z]")

def clean(text: str) -> str:
    return ANSI_RE.sub("", text.replace("\x00", ""))


def open_port(port_name: str) -> serial.Serial:
    ser = serial.Serial(port_name, BAUD, timeout=READ_TIMEOUT)
    wait_for_banner(ser)
    return ser


def wait_for_banner(ser: serial.Serial, max_wait: float = 3.0):
    """After connect/power-up the firmware prints a Usage banner.
    Read until the output goes quiet (or max_wait) so our first command
    isn't swallowed mid-banner."""
    buf = ""
    deadline = time.time() + max_wait
    quiet_since = time.time()
    while time.time() < deadline:
        chunk = ser.read(ser.in_waiting or 1)
        if chunk:
            buf += chunk.decode(errors="replace")
            quiet_since = time.time()
        elif buf and time.time() - quiet_since > 0.4:
            break  # banner finished, line went quiet
    if buf.strip():
        print(f"  [banner] {clean(buf).strip()[:120]}")
    ser.reset_input_buffer()


# what a chip's SUBMENU output contains, per probe_board() observations.
# NOTE: do not check for the chip name itself — the top-level usage line
# ('Usage: gw9121 | gw9241 | gw9251') contains every chip name, so a REJECTED
# select would false-pass on it. 'gw9121' only ever appears in that top-level
# usage, so its presence means we are still at the main menu.
CHIP_SUBMENU_MARKERS = {"gw9241": ("dac",), "gw9251": ("9251",)}


def select_chip(ser: serial.Serial, chip: str):
    """Send the chip-select command (gw9241 / gw9251) and verify it took.

    The board may still be inside a chip submenu from a previous session
    (power not cycled), where 'gw9241'/'gw9251' are invalid commands.
    'exit' returns to the main GW: menu; harmless if already at top level.
    """
    markers = CHIP_SUBMENU_MARKERS[chip]

    def in_submenu(resp: str) -> bool:
        r = resp.lower()
        return any(m in r for m in markers) and "gw9121" not in r

    resp = ""
    for attempt in range(3):
        send(ser, "exit", settle=0.3)  # escape any leftover submenu first
        resp = send(ser, chip, settle=0.3)
        if in_submenu(resp):
            return
        # a successful select may print nothing but the new prompt; probe
        # with an empty line to see where we actually are
        resp = send(ser, "", settle=0.3)
        if in_submenu(resp):
            return
        time.sleep(0.5)
    sys.exit(f"Chip select '{chip}' not acknowledged after 3 tries. "
             f"Last response: {resp!r}\n"
             "If the response shows the WRONG chip's menu, power-cycle both "
             "boards and re-run — auto-identification will re-probe them.")


def send(ser: serial.Serial, cmd: str, settle: float = CMD_DELAY) -> str:
    """Send a command, return everything the board prints back."""
    ser.reset_input_buffer()
    ser.write((cmd + LINE_END).encode())
    ser.flush()
    time.sleep(settle)
    out = []
    deadline = time.time() + READ_TIMEOUT
    while time.time() < deadline:
        chunk = ser.read(ser.in_waiting or 1)
        if chunk:
            out.append(chunk.decode(errors="replace"))
            deadline = time.time() + 0.3  # keep reading while data flows
        else:
            break
    resp = clean("".join(out))
    print(f"  > {cmd}")
    if resp.strip():
        print(f"    {resp.strip()[:200]}")
    return resp


# standalone hex tokens like '0x03', '03', '2F' (word-boundaried, so 'GW9251:'
# or plain words don't produce tokens)
HEX_TOKEN_RE = re.compile(r"\b(?:0x)?([0-9a-fA-F]{1,8})\b")


def _register_match(resp: str, addr: str, expected: str) -> bool:
    """True if some non-echo line of resp reports register addr == expected.

    The exact rr reply format is still unconfirmed on hardware, so accept a
    line that either (a) contains both the address and the expected value as
    standalone hex tokens, or (b) consists of exactly one hex token equal to
    the expected value (value-only reply). Lines containing 'rr' (echo, usage)
    are skipped — a substring check over the whole response would false-pass
    on the echo, the same trap as quirk #4.
    """
    want_addr = int(addr, 16)
    want_val = int(expected, 16)
    for line in re.split(r"[\r\n]+", resp):
        line = line.strip()
        if not line or "rr" in line.lower():
            continue
        toks = [int(t, 16) for t in HEX_TOKEN_RE.findall(line)]
        if not toks:
            continue
        if want_val in toks and (want_addr in toks or len(toks) == 1):
            return True
    return False


def verify_register(ser: serial.Serial, addr: str, expected: str) -> bool:
    """rr <addr> and check the reply reports the expected value."""
    resp = send(ser, f"rr {addr}")
    ok = _register_match(resp, addr, expected)
    if not ok:
        print(f"  !! REGISTER VERIFY FAILED: rr {addr} expected {expected}, got:\n{resp}")
    return ok


# observed 'rd' output: command echo, 'Start conversion', 'No Value', then one
# '<idx> <value>' line per sample (decimal, space-separated, idx from 0).
# idx prefix optional; ':' / '=' separators tolerated. Prompts/echoes don't match.
SAMPLE_LINE_RE = re.compile(
    r"(?:(?P<idx>\d+)\s*[:=\s]\s*)?(?P<val>[+-]?\d+|(?:0x)?[0-9a-fA-F]{4,8})"
)


def _parse_sample_line(line: str) -> int | None:
    m = SAMPLE_LINE_RE.fullmatch(line)
    if not m:
        return None
    val = m.group("val")
    if val.lower().startswith("0x"):
        return int(val, 16)
    if re.fullmatch(r"[+-]?\d+", val):
        return int(val, 10)
    return int(val, 16)  # bare hex with letters, e.g. 'ffab12'


def capture_samples(ser: serial.Serial, n: int = N_SAMPLES) -> list[int]:
    """Send 'rd <n>' and collect exactly n numeric samples.

    Firmware may terminate lines with CR, LF, or CRLF — split on any.
    If the capture falls short, the raw output is dumped to a file so the
    real 'rd' format can be inspected and the parser adjusted.
    """
    ser.reset_input_buffer()
    ser.write(f"rd {n}{LINE_END}".encode())
    ser.flush()

    values: list[int] = []
    raw: list[str] = []  # untouched copy for the debug dump
    buf = ""
    start = time.time()
    hard_deadline = start + CAPTURE_HARD_TIMEOUT
    last_data = start
    reason = "hard timeout"
    max_gap = 0.0  # longest silence observed, for stall diagnostics
    while time.time() < hard_deadline:
        if len(values) >= n:
            reason = "sample count reached"
            break
        chunk = ser.read(ser.in_waiting or 1)
        if not chunk:
            gap = time.time() - last_data
            if gap > CAPTURE_IDLE_TIMEOUT:
                reason = f"stream idle for {gap:.1f}s"
                break
            continue
        max_gap = max(max_gap, time.time() - last_data)
        last_data = time.time()
        text = chunk.decode(errors="replace")
        raw.append(text)
        buf += text.replace("\x00", "")
        # pull complete lines out of the buffer, whatever the line ending
        *lines, buf = re.split(r"[\r\n]+", buf)
        for line in lines:
            v = _parse_sample_line(clean(line).strip())
            if v is not None:
                values.append(v)
    # trailing data without a final newline
    if len(values) < n:
        v = _parse_sample_line(clean(buf).strip())
        if v is not None:
            values.append(v)

    elapsed = time.time() - start
    if len(values) != n:
        rate = len(values) / elapsed if elapsed > 0 else 0.0
        print(f"  !! capture ended: {reason} — {len(values)}/{n} samples "
              f"in {elapsed:.1f}s ({rate:.1f}/s), longest mid-stream gap "
              f"{max_gap:.2f}s")
        dump = Path(f"capture_debug_{datetime.now():%Y%m%d_%H%M%S}.txt")
        dump.write_text(clean("".join(raw)), encoding="utf-8")
        print(f"  !! raw capture saved to {dump}")
    else:
        print(f"  capture complete in {elapsed:.1f}s")
    return values


# ----------------------------- board identification -----------------------------

def list_stlink_ports():
    ports = []
    for p in list_ports.comports():
        # ST-Link VCP is VID 0x0483; keep anything with a serial number as fallback
        if p.vid == 0x0483 or p.serial_number:
            ports.append(p)
    return ports


def probe_board(port_name: str) -> str:
    """Open a port, send an empty line, classify the board by its response.

    DAC board firmware shows a chip-select menu: 'Usage: gw9121 | gw9241 | gw9251'
    (or a gw92xx submenu if left there). The test board boots straight into
    GW9251 mode with a 'GW9251:' prompt and never shows that menu.
    Returns 'dac', 'adc', or 'unknown'.
    """
    try:
        with open_port(port_name) as ser:
            resp = send(ser, "", settle=0.4).lower()
    except serial.SerialException as e:
        print(f"  !! could not probe {port_name}: {e}")
        return "unknown"
    if "gw9121" in resp:            # top-level chip menu only exists on DAC board
        return "dac"
    if "dac" in resp:               # gw9241 submenu (left there from last session)
        return "dac"
    if "9251" in resp:              # GW9251: prompt / submenu
        return "adc"
    return "unknown"


def identify_boards() -> tuple[str, str]:
    """Probe all connected ST-Link ports and return (dac_port, adc_port)."""
    ports = [p.device for p in list_stlink_ports()]
    if len(ports) < 2:
        sys.exit(f"Expected 2 boards, found {len(ports)}. Check USB connections.")

    roles = {port: probe_board(port) for port in ports}
    print("  probe results:", roles)

    dacs = [p for p, r in roles.items() if r == "dac"]
    adcs = [p for p, r in roles.items() if r == "adc"]

    if len(dacs) == 1 and len(adcs) >= 1:
        return dacs[0], adcs[0]

    # ambiguous (e.g. DAC board was left inside a GW9251 submenu):
    # a power cycle resets both boards to a known state
    input("  Could not tell the boards apart. Power-cycle BOTH boards, "
          "then press Enter to re-probe...")
    roles = {port: probe_board(port) for port in ports}
    print("  probe results:", roles)
    dacs = [p for p, r in roles.items() if r == "dac"]
    adcs = [p for p, r in roles.items() if r == "adc"]
    if len(dacs) == 1 and len(adcs) >= 1:
        return dacs[0], adcs[0]

    sys.exit("Board identification failed. Check that both boards are powered "
             "and show their normal prompts in a terminal, then retry.")


# ----------------------------- excel / csv output -----------------------------

def save_csv(values: list[int], dut: int, mode: str) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = Path(f"dut{dut}_{mode}_{ts}.csv")
    path.write_text("\n".join(str(v) for v in values))
    print(f"  saved {path}")
    return path


def _get_sheet(wb):
    if SHEET_NAME:
        return wb[SHEET_NAME]
    return wb.active


def locate_dut_columns(ws, dut: int) -> tuple[int, int]:
    """Scan the header row for 'DUT#<n>' and return (internal_col, channel_a_col).

    Layout: row 5 has merged headers like '2026-07-15 DUT#6' anchored on the
    Internal Short column; Channel A is the column immediately to its right.
    Row 10 is sanity-checked so we never write into the wrong block.
    """
    pattern = re.compile(rf"DUT\s*#\s*{dut}\b")
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if isinstance(v, str) and pattern.search(v):
            mode_int = str(ws.cell(row=MODE_ROW, column=c).value or "")
            mode_cha = str(ws.cell(row=MODE_ROW, column=c + 1).value or "")
            if "Internal" not in mode_int or "Channel" not in mode_cha:
                sys.exit(f"Found DUT#{dut} header at column {c} but mode row {MODE_ROW} "
                         f"doesn't match (got {mode_int!r} / {mode_cha!r}). "
                         "Check the report layout.")
            return c, c + 1
    sys.exit(f"Could not find a 'DUT#{dut}' header in row {HEADER_ROW}. "
             "Add the DUT block to the report or check the number.")


def find_next_empty_dut(ws) -> int | None:
    """Return the lowest DUT number whose Internal Short data area is empty."""
    pattern = re.compile(r"DUT\s*#\s*(\d+)")
    duts = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if isinstance(v, str):
            m = pattern.search(v)
            if m:
                duts.append((int(m.group(1)), c))
    for dut, c in sorted(duts):
        if ws.cell(row=DATA_START_ROW, column=c).value is None:
            return dut
    return None


def write_excel(values: list[int], dut: int, mode: str, excel_path: Path):
    from openpyxl import load_workbook

    if not excel_path.exists():
        print(f"  !! {excel_path} not found, skipping Excel write (CSV is saved).")
        return

    wb = load_workbook(excel_path)  # default load: formulas preserved
    ws = _get_sheet(wb)
    col_int, col_cha = locate_dut_columns(ws, dut)
    col = col_int if mode == "internal" else col_cha

    existing = ws.cell(row=DATA_START_ROW, column=col).value
    if existing is not None:
        ans = input(f"  DUT#{dut} {mode} column already has data "
                    f"(row {DATA_START_ROW} = {existing}). Overwrite? [y/N] ").strip().lower()
        if ans != "y":
            print("  skipped Excel write (CSV is saved).")
            return

    for i, v in enumerate(values):
        ws.cell(row=DATA_START_ROW + i, column=col, value=v)
    wb.save(excel_path)
    from openpyxl.utils import get_column_letter
    print(f"  wrote {len(values)} samples to {excel_path} "
          f"[{ws.title}!{get_column_letter(col)}{DATA_START_ROW}:"
          f"{get_column_letter(col)}{DATA_START_ROW + len(values) - 1}]")


# ----------------------------- test sequences -----------------------------

def setup_dac(dac: serial.Serial):
    print("\n[DAC board setup]")
    select_chip(dac, "gw9241")
    send(dac, "dac init", settle=0.5)
    send(dac, f"dac set 1 {DAC_LOW}")
    send(dac, f"dac set 2 {DAC_HIGH}")


def run_mode(adc: serial.Serial, dut: int, mode: str, reg04: str, excel_path: Path):
    label = "Internal Short" if mode == "internal" else "Channel A"
    print(f"\n[{label}] DUT #{dut}")

    send(adc, "wr 0x03 0x02")
    if not verify_register(adc, "0x03", "0x02"):
        input("  Register 0x03 mismatch. Fix and press Enter to retry, Ctrl+C to abort...")
        send(adc, "wr 0x03 0x02")
        if not verify_register(adc, "0x03", "0x02"):
            sys.exit("Register 0x03 verify failed twice. Aborting.")

    send(adc, f"wr 0x04 {reg04}")
    if not verify_register(adc, "0x04", reg04):
        input(f"  Register 0x04 mismatch. Fix and press Enter to retry, Ctrl+C to abort...")
        send(adc, f"wr 0x04 {reg04}")
        if not verify_register(adc, "0x04", reg04):
            sys.exit("Register 0x04 verify failed twice. Aborting.")

    print(f"  capturing {N_SAMPLES} samples...")
    values = capture_samples(adc)
    if len(values) != N_SAMPLES:
        print(f"  !! Only got {len(values)}/{N_SAMPLES} samples.")
        retry = input("  Retry capture? [y/N] ").strip().lower()
        if retry == "y":
            values = capture_samples(adc)
        if len(values) != N_SAMPLES:
            sys.exit(f"Capture incomplete ({len(values)}/{N_SAMPLES}). Aborting.")

    save_csv(values, dut, mode)
    write_excel(values, dut, mode, excel_path)


def wait_for_boards(poll: float = 1.0) -> tuple[str, str]:
    """Block until two boards enumerate on USB after a power cycle,
    then identify which is which by probing their firmware prompts."""
    print("  waiting for both boards on USB", end="", flush=True)
    while True:
        ports = list_stlink_ports()
        if len(ports) >= 2:
            print()
            time.sleep(1.5)  # let Windows finish enumerating before opening
            dac_port, adc_port = identify_boards()
            print(f"  boards identified: DAC={dac_port}  ADC={adc_port}")
            return dac_port, adc_port
        print(".", end="", flush=True)
        time.sleep(poll)


def verify_adc_ready(adc: serial.Serial):
    """The test board boots directly into GW9251 mode (no chip-select menu).
    Just confirm the GW9251 prompt responds; try 'gw9251' once as fallback."""
    resp = send(adc, "", settle=0.4)
    if "9251" in resp.lower():
        return
    resp = send(adc, "gw9251", settle=0.4)
    if "9251" in resp.lower():
        return
    sys.exit(f"Test board is not responding with a GW9251 prompt. "
             f"Last response: {resp!r}\nPower-cycle the board and retry.")


def test_one_dut(dac_port: str, adc_port: str, dut: int, excel_path: Path):
    """Full cycle for one chip: DAC init/set + Internal Short + Channel A."""
    with open_port(dac_port) as dac, open_port(adc_port) as adc:
        verify_adc_ready(adc)
        setup_dac(dac)  # power was cycled, so DAC state is gone: always re-init

        run_mode(adc, dut, "internal", "0x60", excel_path)   # noise floor
        run_mode(adc, dut, "channel_a", "0x00", excel_path)  # accuracy


def next_dut_from_excel(excel_path: Path) -> int | None:
    if not excel_path.exists():
        return None
    from openpyxl import load_workbook
    wb = load_workbook(excel_path, read_only=True)
    ws = _get_sheet(wb)
    dut = find_next_empty_dut(ws)
    wb.close()
    return dut


def main():
    global SHEET_NAME
    ap = argparse.ArgumentParser()
    ap.add_argument("--dut", type=int, default=None,
                    help="starting DUT number (default: first empty DUT block in the report)")
    ap.add_argument("--excel", type=Path, default=EXCEL_PATH,
                    help=f"path to the Excel report (default: {EXCEL_PATH})")
    ap.add_argument("--sheet", type=str, default=None,
                    help="sheet name (default: active sheet)")
    args = ap.parse_args()
    if args.sheet:
        SHEET_NAME = args.sheet

    if not args.excel.exists():
        print(f"Note: {args.excel} not found. CSVs will still be saved.")

    dut = args.dut
    if dut is None:
        dut = next_dut_from_excel(args.excel)
        if dut is None:
            sys.exit(f"Could not auto-detect next DUT from {args.excel}. "
                     "Pass --dut explicitly.")
        print(f"Auto-detected next empty DUT block: DUT#{dut}")

    while True:
        dac_port, adc_port = wait_for_boards()
        test_one_dut(dac_port, adc_port, dut, args.excel)

        print(f"\nDUT #{dut} complete.")
        ans = input("Power off, swap chip, power on, then press Enter for next DUT "
                    "(or type q to quit): ").strip().lower()
        if ans == "q":
            break
        dut += 1
        # user has cut power by now; loop back to wait_for_boards, which
        # blocks until both boards re-appear, then DAC is re-initialized
        # automatically inside test_one_dut.

    print("\nAll done.")


if __name__ == "__main__":
    main()

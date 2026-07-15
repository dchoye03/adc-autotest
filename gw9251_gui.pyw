"""
GW9251 ADC test — GUI front-end (add-on).

Imports gw9251_autotest and reuses its serial/Excel functions unchanged.
Lets you pick the DAC voltage codes, the PGA (reg 0x03) value, and which
mode(s) to run (Internal Short / Channel A), then runs one full DUT cycle.

Launch by double-clicking the desktop shortcut (runs with pythonw, no
console — all output appears in the log pane below).
"""

import builtins
import math
import os
import queue
import re
import statistics
import sys
import threading
import traceback
from copy import copy
from datetime import datetime
from pathlib import Path

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkinter.scrolledtext import ScrolledText

# make the import work no matter where the shortcut starts us
sys.path.insert(0, str(Path(__file__).resolve().parent))
import gw9251_autotest as g

DUT_RE = re.compile(r"DUT\s*#\s*(\d+)")

PGA_GAIN = 64  # report row 8 'PGA (V/V)'; fixed while reg 0x03 = 0x02


# ---------------- result summary (mirrors the Excel report formulas) ----------------

def summarize_internal(values: list[int]):
    """Internal Short: (std, enob) — report rows 12 and 11.
    STD = STDEV.S(samples); ENOB = LOG(0.9*2^24/STD, 2)."""
    if len(values) < 2:
        return None, None
    std = statistics.stdev(values)
    enob = math.log2(0.9 * 2 ** 24 / std) if std > 0 else None
    return std, enob


def summarize_channel_a(values: list[int], dac1: int, dac2: int):
    """Channel A: (avg, vin_calc, vin_actual, accuracy) — report rows 13/15/16/20.
    Vin(calc) = 3/2^23 * AVG / PGA;  Vin(actual) = (dac2-dac1) DAC codes in µV;
    Accuracy = calc/actual."""
    if not values:
        return None, None, None, None
    avg = statistics.fmean(values)
    vin_calc = 3 / 2 ** 23 * avg / PGA_GAIN
    vin_actual = (dac2 - dac1) / 1e6
    accuracy = vin_calc / vin_actual if vin_actual else None
    return avg, vin_calc, vin_actual, accuracy

# (key, checkbox label, reg 0x04 value) — keys match gw9251_autotest's
# mode strings so save_csv/write_excel column selection works unchanged
MODES = [
    ("internal",  "Internal Short (0x04 = 0x60)", "0x60"),
    ("channel_a", "Channel A (0x04 = 0x00)",      "0x00"),
]
MODE_LABELS = {key: label for key, label, _ in MODES}


# ---------------- Excel: auto-extend DUT blocks (add-on) ----------------

def _scan_duts(ws) -> dict[int, int]:
    """Map DUT number -> anchor column, from the header row."""
    duts = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=g.HEADER_ROW, column=c).value
        if isinstance(v, str):
            m = DUT_RE.search(v)
            if m:
                duts[int(m.group(1))] = c
    return duts


def _copy_block(ws, src_col: int, dst_col: int, new_dut: int):
    """Clone one 2-column DUT block's template (header, mode row, formulas,
    styles, merges) from src_col to dst_col. Data rows stay empty (styles
    only); formulas are re-pointed at the new columns."""
    from openpyxl.cell.cell import MergedCell
    from openpyxl.formula.translate import Translator
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.cell_range import CellRange

    # replicate merged ranges from the source block FIRST (skipping ones the
    # report already pre-merged, e.g. the empty W5:X5..AK5:AL5 header cells),
    # so value writes below can target the anchors of the final merge layout
    shift = dst_col - src_col
    for rng in list(ws.merged_cells.ranges):
        if (rng.min_col >= src_col and rng.max_col <= src_col + 1
                and rng.min_row >= g.HEADER_ROW
                and rng.max_row < g.DATA_START_ROW):
            new_rng = CellRange(min_col=rng.min_col + shift, min_row=rng.min_row,
                                max_col=rng.max_col + shift, max_row=rng.max_row)
            if not any(new_rng.issubset(ex) or ex.issubset(new_rng)
                       for ex in ws.merged_cells.ranges):
                ws.merge_cells(new_rng.coord)

    last_data_row = g.DATA_START_ROW + g.N_SAMPLES - 1
    for off in (0, 1):
        sc, dc = src_col + off, dst_col + off
        src_dim = ws.column_dimensions[get_column_letter(sc)]
        ws.column_dimensions[get_column_letter(dc)].width = src_dim.width
        for r in range(g.HEADER_ROW, last_data_row + 1):
            s = ws.cell(row=r, column=sc)
            d = ws.cell(row=r, column=dc)
            d._style = copy(s._style)
            if r >= g.DATA_START_ROW or isinstance(d, MergedCell):
                continue  # sample area / merged non-anchor: style only
            v = s.value
            if isinstance(v, str) and v.startswith("="):
                d.value = Translator(v, origin=s.coordinate).translate_formula(
                    d.coordinate)
            elif r == g.HEADER_ROW and isinstance(v, str):
                d.value = DUT_RE.sub(f"DUT#{new_dut}", v)
            elif not isinstance(s, MergedCell):
                d.value = v


class QueueWriter:
    """File-like stdout replacement: forwards prints to the GUI log queue,
    optionally mirroring into a list (used for the per-run log file)."""

    def __init__(self, q: queue.Queue, mirror: list | None = None):
        self.q = q
        self.mirror = mirror

    def write(self, s: str):
        if s:
            self.q.put(s)
            if self.mirror is not None:
                self.mirror.append(s)

    def flush(self):
        pass


class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        root.title("GW9251 ADC Test")
        root.minsize(560, 480)
        self.q: queue.Queue[str] = queue.Queue()

        frm = ttk.Frame(root, padding=10)
        frm.grid(sticky="nsew")
        root.columnconfigure(0, weight=1)
        root.rowconfigure(0, weight=1)

        self._row = 0

        # -------- prerequisites notice --------
        notice = ttk.LabelFrame(frm, text="준비사항 (read me first!)", padding=6)
        notice.grid(row=self._row, column=0, columnspan=2, sticky="we", pady=(0, 8))
        ttk.Label(notice, justify="left", text=(
            "1) 엑셀 파일과 파이썬 파일을 같은 폴더에 두세요\n"
            "    (gw9251_gui.pyw + gw9251_autotest.py + 엑셀 리포트)\n"
            "2) 테스트 중에는 엑셀 파일을 열어두지 마세요 (저장이 안 됩니다)\n"
            "3) Tera Term은 꼭 닫아 주세요 (COM 포트 독점)\n"
            "4) 보드 2개 USB 연결 + 전원 ON\n"
            "5) 최초 1회만: 라이브러리 설치\n"
            "    윈도우 키 누르고 cmd 입력 → Enter (검은 창이 열림)\n"
            "    그 창에 아래를 입력하고 Enter:\n"
            "    pip install pyserial openpyxl"
        )).grid(sticky="w")
        self._row += 1

        def add_entry(label: str, default: str) -> tk.StringVar:
            ttk.Label(frm, text=label).grid(row=self._row, column=0, sticky="w")
            var = tk.StringVar(value=default)
            ttk.Entry(frm, textvariable=var, width=20).grid(
                row=self._row, column=1, sticky="we", pady=2)
            self._row += 1
            return var

        self.dac1 = add_entry("Voltage: dac set 1 (AINN) code", str(g.DAC_LOW))
        self.dac2 = add_entry("Voltage: dac set 2 (AINP) code", str(g.DAC_HIGH))
        self.nsamp = add_entry("How many samples", str(g.N_SAMPLES))
        self.dut = add_entry("DUT #  (blank = next empty in Excel)", "")
        # Excel report: dropdown of .xlsx files in this folder + Browse…
        ttk.Label(frm, text="Excel report").grid(row=self._row, column=0, sticky="w")
        self.excel = tk.StringVar(value="26.7.14.adc.xlsx")
        self.excel_box = ttk.Combobox(frm, textvariable=self.excel, width=24)
        self.excel_box.grid(row=self._row, column=1, sticky="we", pady=2)

        def refresh_xlsx_list():
            here = Path(__file__).resolve().parent
            names = sorted(p.name for p in here.glob("*.xlsx")
                           if not p.name.startswith("~$"))  # skip Excel lock files
            self.excel_box["values"] = names

        self.excel_box.configure(postcommand=refresh_xlsx_list)
        refresh_xlsx_list()
        if self.excel.get() not in self.excel_box["values"] and self.excel_box["values"]:
            self.excel.set(self.excel_box["values"][0])

        def browse_xlsx():
            path = filedialog.askopenfilename(
                title="엑셀 리포트 선택",
                initialdir=Path(__file__).resolve().parent,
                filetypes=[("Excel 파일", "*.xlsx")])
            if path:
                self.excel.set(path)

        ttk.Button(frm, text="찾아보기…", command=browse_xlsx, width=10).grid(
            row=self._row, column=2, padx=(4, 0))
        self._row += 1

        self.mode_vars: dict[str, tk.BooleanVar] = {}
        for key, label, _ in MODES:
            var = tk.BooleanVar(value=True)
            ttk.Checkbutton(frm, text=label, variable=var).grid(
                row=self._row, column=0, columnspan=2, sticky="w")
            self.mode_vars[key] = var
            self._row += 1

        self.run_btn = ttk.Button(frm, text="Run test", command=self.run_clicked)
        self.run_btn.grid(row=self._row, column=0, columnspan=2, pady=6, sticky="we")
        self._row += 1

        self.log_box = ScrolledText(frm, width=80, height=22, state="disabled",
                                    font=("Consolas", 9))
        self.log_box.grid(row=self._row, column=0, columnspan=2, sticky="nsew")
        frm.rowconfigure(self._row, weight=1)
        frm.columnconfigure(1, weight=1)

        root.after(100, self._poll_log)

    # ---------------- log plumbing ----------------

    def _poll_log(self):
        try:
            while True:
                s = self.q.get_nowait()
                self.log_box.configure(state="normal")
                self.log_box.insert("end", s)
                self.log_box.see("end")
                self.log_box.configure(state="disabled")
        except queue.Empty:
            pass
        self.root.after(100, self._poll_log)

    def gui_input(self, prompt: str = "") -> str:
        """Stand-in for builtins.input() while a test runs (worker thread).
        '[y/N]' prompts become Yes/No dialogs; 'press Enter' prompts an OK box."""
        self.q.put(prompt + "\n")
        done = threading.Event()
        result = {"v": ""}

        def ask():
            if "y/n" in prompt.lower():
                result["v"] = "y" if messagebox.askyesno("GW9251", prompt) else "n"
            else:
                messagebox.showinfo("GW9251", prompt or "Continue?")
            done.set()

        self.root.after(0, ask)
        done.wait()
        self.q.put(f"  -> {result['v'] or '(ok)'}\n")
        return result["v"]

    # ---------------- run ----------------

    def validate(self) -> dict:
        def code(s: str, name: str) -> int:
            try:
                return int(s.strip())
            except ValueError:
                raise ValueError(f"{name} must be an integer DAC code")

        dac1 = code(self.dac1.get(), "dac set 1")
        dac2 = code(self.dac2.get(), "dac set 2")

        try:
            n = int(self.nsamp.get().strip())
        except ValueError:
            raise ValueError("'How many samples' must be a whole number")
        if n <= 0:
            raise ValueError("'How many samples' must be at least 1")

        modes = [(key, reg) for key, _, reg in MODES if self.mode_vars[key].get()]
        if not modes:
            raise ValueError("Select at least one mode")

        dut_s = self.dut.get().strip()
        dut = int(dut_s) if dut_s else None
        excel = Path(self.excel.get().strip())
        return dict(dac1=dac1, dac2=dac2, n=n, modes=modes, dut=dut,
                    excel=excel)

    def run_clicked(self):
        try:
            params = self.validate()
        except ValueError as e:
            messagebox.showerror("GW9251", str(e))
            return
        # the Excel report layout (rows 22-1045 + ENOB/STD formulas) assumes
        # exactly N_SAMPLES rows — warn before writing a different count
        if params["n"] != g.N_SAMPLES and not messagebox.askyesno(
                "GW9251",
                f"엑셀 리포트 수식은 샘플 {g.N_SAMPLES}개 기준입니다.\n"
                f"{params['n']}개로 진행하면 ENOB/STD 수식이 맞지 않을 수 "
                "있어요.\n계속할까요?"):
            return
        if not params["excel"].exists() and not messagebox.askyesno(
                "GW9251",
                f"엑셀 파일이 없어요: {params['excel']}\n"
                "(엑셀 파일을 이 폴더에 같이 두세요)\n"
                "CSV만 저장하고 계속할까요?"):
            return
        self.run_btn.configure(state="disabled", text="Running…  (see log)")
        threading.Thread(target=self._worker, args=(params,), daemon=True).start()

    def _worker(self, p: dict):
        start_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        run_log: list[str] = []
        old = (sys.stdout, sys.stderr, builtins.input)
        sys.stdout = sys.stderr = QueueWriter(self.q, run_log)
        builtins.input = self.gui_input  # console prompts -> dialogs
        try:
            print(f"=== run {start_ts}  dac1={p['dac1']} dac2={p['dac2']} "
                  f"n={p['n']} excel={p['excel']} ===")
            dut, summary = self._run_test(p)
            print("\n=== DONE ===")
            if p["excel"].exists():
                print(f"  opening {p['excel']} ...")
                os.startfile(p["excel"].resolve())
            result_txt = ("\n".join(summary) + "\n\n") if summary else ""
            self.root.after(0, lambda: messagebox.showinfo(
                "GW9251",
                f"DUT #{dut} 완료!\n\n" + result_txt +
                "다음 칩 테스트 순서:\n"
                "  1) 보드 전원 OFF\n"
                "  2) 칩 교체 (새 칩 장착)\n"
                "  3) 전원 ON\n"
                "  4) Run test 다시 클릭\n\n"
                "※ 자동으로 열린 엑셀은 다음 테스트 전에 닫아 주세요\n"
                "   (열려 있으면 저장할 때 다시 물어봅니다)"))
        except SystemExit as e:
            print(f"\n!! aborted: {e}")
        except Exception:
            traceback.print_exc()
        finally:
            sys.stdout, sys.stderr, builtins.input = old
            try:
                log_path = Path(f"run_{start_ts}.log")
                log_path.write_text("".join(run_log), encoding="utf-8")
                self.q.put(f"  run log saved: {log_path}\n")
            except OSError as e:
                self.q.put(f"  !! could not save run log: {e}\n")
            self.root.after(0, lambda: self.run_btn.configure(
                state="normal", text="Run test"))

    def _load_report(self, excel: Path):
        """load_workbook with the 'Excel has the file open' retry dialog."""
        from openpyxl import load_workbook
        while True:
            try:
                return load_workbook(excel)
            except PermissionError:
                input("엑셀 파일이 열려 있어요! Excel을 닫은 다음 OK를 눌러 주세요.")

    def _ensure_dut_block(self, excel: Path, dut: int):
        """If the report has no DUT#<dut> block, clone the last block's
        template into new columns (creating every block up to <dut>)."""
        wb = self._load_report(excel)
        ws = wb[g.SHEET_NAME] if g.SHEET_NAME else wb.active
        duts = _scan_duts(ws)
        if not duts:
            raise SystemExit(f"row {g.HEADER_ROW}에서 DUT# 헤더를 못 찾았어요 — "
                             "리포트 레이아웃 확인 필요")
        if dut in duts:
            wb.close()
            return
        last = max(duts)
        if dut < last:
            raise SystemExit(f"DUT#{dut} 블록이 없는데 리포트에는 DUT#{last}까지 "
                             "있어요 — DUT 번호를 확인해 주세요")
        last_data_row = g.DATA_START_ROW + g.N_SAMPLES - 1
        for nd in range(last + 1, dut + 1):
            dst = duts[last] + 2 * (nd - last)
            _copy_block(ws, duts[last], dst, nd)
            # the report keeps scratch data right of the DUT blocks (e.g.
            # cols Y/Z) — a new block may land on it. Never delete silently.
            leftover = sum(
                1 for r in range(g.DATA_START_ROW, last_data_row + 1)
                for c in (dst, dst + 1)
                if ws.cell(row=r, column=c).value is not None)
            if leftover:
                if input(f"새 DUT#{nd} 블록 자리에 기존 값 {leftover}개가 "
                         "있어요 (스크래치 데이터?).\n지우고 계속할까요? "
                         "[y/N] ").strip().lower() != "y":
                    raise SystemExit(f"DUT#{nd} 블록 생성 취소 — 리포트의 "
                                     "스크래치 데이터를 옮긴 후 다시 실행해 "
                                     "주세요")
                for r in range(g.DATA_START_ROW, last_data_row + 1):
                    for c in (dst, dst + 1):
                        ws.cell(row=r, column=c).value = None
            print(f"  리포트에 DUT#{nd} 블록을 새로 만들었어요 (템플릿 복사)")
        while True:
            try:
                wb.save(excel)
                return
            except PermissionError:
                input("엑셀 파일이 열려 있어요! Excel을 닫은 다음 OK를 눌러 주세요.")

    def _run_test(self, p: dict):
        dut = p["dut"]
        if dut is None:
            if not p["excel"].exists():
                raise SystemExit("엑셀 파일이 없으면 DUT 자동감지가 안 돼요 — "
                                 "DUT #를 직접 입력해 주세요")
            dut = g.next_dut_from_excel(p["excel"])
            if dut is None:
                # every existing block already has data -> extend the report
                wb = self._load_report(p["excel"])
                ws = wb[g.SHEET_NAME] if g.SHEET_NAME else wb.active
                duts = _scan_duts(ws)
                wb.close()
                if not duts:
                    raise SystemExit("리포트에서 DUT# 헤더를 못 찾았어요")
                dut = max(duts) + 1
                print(f"모든 DUT 블록이 가득 참 → 새 블록 DUT#{dut} 생성")
            else:
                print(f"auto-detected next empty DUT block: DUT#{dut}")
        if p["excel"].exists():
            self._ensure_dut_block(p["excel"], dut)
        summary: list[str] = []

        dac_port, adc_port = g.wait_for_boards()
        with g.open_port(dac_port) as dac, g.open_port(adc_port) as adc:
            g.verify_adc_ready(adc)

            print("\n[DAC board setup]")
            g.select_chip(dac, "gw9241")
            g.send(dac, "dac init", settle=0.5)
            g.send(dac, f"dac set 1 {p['dac1']}")
            g.send(dac, f"dac set 2 {p['dac2']}")

            n = p["n"]
            for mode, reg04 in p["modes"]:
                print(f"\n[{MODE_LABELS[mode]}] DUT #{dut}")
                for addr, val in (("0x03", "0x02"), ("0x04", reg04)):
                    g.send(adc, f"wr {addr} {val}")
                    if not g.verify_register(adc, addr, val):
                        input(f"Register {addr} mismatch — press OK to retry once")
                        g.send(adc, f"wr {addr} {val}")
                        if not g.verify_register(adc, addr, val):
                            raise SystemExit(f"register {addr} verify failed twice")

                print(f"  capturing {n} samples...")
                values = g.capture_samples(adc, n)
                if len(values) != n:
                    if input(f"Only {len(values)}/{n} samples. "
                             "Retry capture? [y/N] ").lower() == "y":
                        values = g.capture_samples(adc, n)
                if len(values) != n:
                    raise SystemExit(f"capture incomplete ({len(values)}/{n})")

                g.save_csv(values, dut, mode)
                self._write_excel_retry(values, dut, mode, p["excel"])

                # live result summary, same math as the Excel report formulas
                if mode == "internal":
                    std, enob = summarize_internal(values)
                    if std is not None:
                        line = (f"Internal Short: STD = {std:,.2f}, "
                                f"ENOB = {enob:.2f} bit" if enob is not None
                                else f"Internal Short: STD = {std:,.2f}")
                        print(f"  >> {line}")
                        summary.append(line)
                else:
                    avg, vin_c, vin_a, acc = summarize_channel_a(
                        values, p["dac1"], p["dac2"])
                    if avg is not None:
                        line = (f"Channel A: 평균 = {avg:,.0f}, "
                                f"Vin(calc) = {vin_c * 1000:.3f} mV / "
                                f"{vin_a * 1000:.1f} mV")
                        if acc is not None:
                            line += f", 정확도 = {acc * 100:.3f} %"
                        print(f"  >> {line}")
                        summary.append(line)
        return dut, summary

    def _write_excel_retry(self, values, dut, mode, excel):
        """write_excel, but if the report is open in Excel (PermissionError on
        load/save) ask the user to close it and retry instead of crashing."""
        while True:
            try:
                g.write_excel(values, dut, mode, excel)
                return
            except PermissionError:
                input("엑셀 파일이 열려 있어서 저장할 수 없어요!\n"
                      "Excel을 닫은 다음 OK를 눌러 주세요 — 다시 저장합니다.")


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()

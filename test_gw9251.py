"""
Regression tests for gw9251_autotest.py + gw9251_gui.pyw.
No hardware needed. Run:  python test_gw9251.py
"""

import builtins
import importlib.util
import math
import os
import shutil
import statistics
import sys
import tempfile
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))

# console-safe output regardless of codepage (arrows, Korean, etc.)
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import gw9251_autotest as g

CHECKS = 0


def ok(name, cond):
    global CHECKS
    assert cond, f"FAIL: {name}"
    CHECKS += 1
    print(f"  ok - {name}")


def load_gui():
    spec = importlib.util.spec_from_file_location("gw9251_gui",
                                                  HERE / "gw9251_gui.pyw")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


# ---------------- clean() / ANSI stripping ----------------

def test_clean():
    ok("NUL bytes stripped", g.clean("a\x00b\x00c") == "abc")
    ok("standard ANSI stripped", g.clean("\x1b[2Jhello\x1b[0m") == "hello")
    # firmware emits non-standard ESC[-1D cursor moves (quirk from real dump)
    echo = "r\x1b[-1Dd\x1b[-1D \x1b[-1D1\x1b[-1D0\x1b[-1D2\x1b[-1D4\x1b[-1D"
    ok("non-standard ESC[-1D stripped", g.clean(echo) == "rd 1024")


# ---------------- sample line parser ----------------

def test_parse_sample_line():
    cases = {
        "0 -6": -6,                 # real rd format: '<idx> <value>'
        "1023 -128": -128,
        "12: 16734567": 16734567,
        "12 = 0xABCDEF": 0xABCDEF,
        "16734567": 16734567,       # bare value
        "-34567": -34567,
        "0x00FFAB12": 0x00FFAB12,
        "ffab12": 0xFFAB12,
        "Start conversion": None,   # junk must not parse
        "No Value": None,
        "GW9251:": None,
        "rd 1024": None,
        "Usage: set | wr | rr | rd": None,
        "": None,
    }
    for line, want in cases.items():
        got = g._parse_sample_line(line)
        ok(f"parse {line!r} -> {want}", got == want)


def test_capture_replay():
    """Replay the real capture dump through the parsing path."""
    dump = HERE / "capture_debug_20260715_113713.txt"
    if not dump.exists():
        print("  skip - capture dump not present")
        return
    import re
    values = []
    for line in re.split(r"[\r\n]+", dump.read_text(encoding="utf-8")):
        v = g._parse_sample_line(g.clean(line).strip())
        if v is not None:
            values.append(v)
    ok("replay: 341 samples parsed from real dump", len(values) == 341)
    ok("replay: first samples correct", values[:3] == [-6, -9, -17])


# ---------------- register verify matching ----------------

def test_register_match():
    m = g._register_match
    ok("'0x03 0x02' matches", m("0x03 0x02", "0x03", "0x02"))
    ok("'Reg[0x03] = 0x02' matches", m("Reg[0x03] = 0x02", "0x03", "0x02"))
    ok("'03 02' matches", m("03 02", "0x03", "0x02"))
    ok("value-only '0x02' matches", m("0x02", "0x03", "0x02"))
    ok("value on second line matches", m("rr 0x03\r\n0x02\r\n", "0x03", "0x02"))
    ok("wrong value fails", not m("0x03 0x04", "0x03", "0x02"))
    ok("echo alone fails (addr==value trap)", not m("rr 0x03", "0x03", "0x03"))
    ok("prompt alone fails", not m("GW9251:", "0x03", "0x02"))
    ok("usage line fails", not m("set | wr | rr | rd | cal", "0x03", "0x02"))
    ok("empty fails", not m("", "0x03", "0x02"))


# ---------------- GUI: summaries (mirror Excel formulas) ----------------

def test_summaries(gui):
    vals = [-1, 1] * 512
    std, enob = gui.summarize_internal(vals)
    ok("internal STD == statistics.stdev", abs(std - statistics.stdev(vals)) < 1e-12)
    ok("ENOB matches report formula", abs(enob - math.log2(0.9 * 2**24 / std)) < 1e-12)
    ok("ENOB near 23.85 for std~1", abs(enob - 23.8489) < 0.01)
    std0, enob0 = gui.summarize_internal([5, 5, 5])
    ok("zero-STD gives ENOB None", std0 == 0 and enob0 is None)
    ok("too-few samples gives None", gui.summarize_internal([1]) == (None, None))

    avg, vin_c, vin_a, acc = gui.summarize_channel_a([7158279] * 4,
                                                     1480000, 1520000)
    ok("channel A avg", avg == 7158279)
    ok("vin_actual = 40 mV from DAC codes", abs(vin_a - 0.04) < 1e-12)
    ok("accuracy ~100% at nominal avg", abs(acc - 1.0) < 1e-5)
    ok("empty channel A gives Nones",
       gui.summarize_channel_a([], 0, 0) == (None, None, None, None))


# ---------------- GUI: validation ----------------

def test_gui_validation(gui, app):
    p = app.validate()
    ok("default DAC codes", p["dac1"] == 1480000 and p["dac2"] == 1520000)
    ok("default n = N_SAMPLES", p["n"] == g.N_SAMPLES)
    ok("default modes = both",
       p["modes"] == [("internal", "0x60"), ("channel_a", "0x00")])

    app.nsamp.set("256")
    ok("custom n accepted", app.validate()["n"] == 256)
    for bad in ("0", "-5", "abc", ""):
        app.nsamp.set(bad)
        try:
            app.validate()
            raise AssertionError(f"bad n accepted: {bad!r}")
        except ValueError:
            ok(f"bad n rejected: {bad!r}", True)
    app.nsamp.set(str(g.N_SAMPLES))

    app.dut.set("7")
    ok("dut parsed", app.validate()["dut"] == 7)
    app.dut.set("")

    app.mode_vars["internal"].set(False)
    app.mode_vars["channel_a"].set(False)
    try:
        app.validate()
        raise AssertionError("no-mode accepted")
    except ValueError:
        ok("no-mode rejected", True)
    app.mode_vars["internal"].set(True)
    app.mode_vars["channel_a"].set(True)


# ---------------- GUI: excel write retry on PermissionError ----------------

def test_excel_retry(gui, app):
    calls = {"write": 0, "ask": 0}

    def fake_write(values, dut, mode, excel):
        calls["write"] += 1
        if calls["write"] == 1:
            raise PermissionError("locked by Excel")

    def fake_input(prompt=""):
        calls["ask"] += 1
        return ""

    real_write, real_input = gui.g.write_excel, builtins.input
    gui.g.write_excel, builtins.input = fake_write, fake_input
    try:
        app._write_excel_retry([1, 2], 99, "internal", Path("x.xlsx"))
    finally:
        gui.g.write_excel, builtins.input = real_write, real_input
    ok("excel retry: locked once -> asked once, wrote twice",
       calls == {"write": 2, "ask": 1})


# ---------------- GUI: DUT block auto-extend ----------------

def test_block_extend(gui, app, report: Path):
    with tempfile.TemporaryDirectory() as td:
        test = Path(td) / "extend.xlsx"
        shutil.copy(report, test)

        prompts = []

        def yes_input(prompt=""):
            prompts.append(prompt)
            return "y"

        real_input = builtins.input
        builtins.input = yes_input
        try:
            from openpyxl import load_workbook
            wb = load_workbook(test, read_only=False)
            last = max(gui._scan_duts(wb.active))
            wb.close()
            app._ensure_dut_block(test, last + 2)  # creates 2 new blocks
        finally:
            builtins.input = real_input

        wb = load_workbook(test)
        ws = wb.active
        for nd in (last + 1, last + 2):
            ci, ca = g.locate_dut_columns(ws, nd)  # strict layout cross-check
            ok(f"clone DUT#{nd}: locator accepts block", True)
            ok(f"clone DUT#{nd}: data area empty",
               ws.cell(row=g.DATA_START_ROW, column=ci).value is None
               and ws.cell(row=g.DATA_START_ROW, column=ca).value is None)
            from openpyxl.utils import get_column_letter
            f11 = ws.cell(row=11, column=ci).value
            ok(f"clone DUT#{nd}: ENOB formula re-pointed",
               isinstance(f11, str) and f"{get_column_letter(ci)}12" in f11)
        # idempotent: existing block -> no rewrite
        mtime = os.path.getmtime(test)
        app._ensure_dut_block(test, last + 2)
        ok("extend is idempotent (no rewrite)", os.path.getmtime(test) == mtime)


# ---------------- main ----------------

def main():
    print("gw9251_autotest tests:")
    test_clean()
    test_parse_sample_line()
    test_capture_replay()
    test_register_match()

    print("gw9251_gui tests:")
    gui = load_gui()
    import tkinter as tk
    root = tk.Tk()
    root.withdraw()
    app = gui.App(root)
    try:
        test_summaries(gui)
        test_gui_validation(gui, app)
        test_excel_retry(gui, app)
        report = HERE / "test_copy.xlsx"
        if not report.exists():
            report = HERE / "26.7.14.adc.xlsx"
        if report.exists():
            test_block_extend(gui, app, report)
        else:
            print("  skip - no report file for block-extend test")
    finally:
        root.destroy()

    print(f"\nALL {CHECKS} CHECKS PASSED")


if __name__ == "__main__":
    main()
